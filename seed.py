"""
seed.py – Roda UMA VEZ para:
  1) Criar o banco dados.db
  2) Importar os PJs do Excel para a tabela pj_contratos
  3) Criar o usuário admin (senha: admin123 — troque depois!)

Como usar:
  .\\venv\\Scripts\\python.exe seed.py
"""
import os
import sys
import openpyxl
from datetime import datetime

# Bootstrap Flask app context
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
from app import app
from models import db, User, PjContrato, CltRegistro, AgrupadorCC

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'PJ - CONTRATOS REAJUSTADOS.xlsx')
CLT_EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Base.xlsx')
DEPARA_EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'depara.xlsx')
ADMIN_USER = 'admin'
ADMIN_PASS = 'admin123'  # ⚠️ Troque para uma senha segura após o primeiro login!

# Mapeamento: coluna do Excel → campo do modelo
# Linha 3 do Excel é o cabeçalho; colunas de datas começam na coluna 12 (L)
MONTH_COL_MAP = {
    12: 'valor_jan_2026',
    13: 'valor_fev_2026',
    14: 'valor_mar_2026',
    15: 'valor_abr_2026',
    16: 'valor_mai_2026',
    17: 'valor_jun_2026',
    18: 'valor_jul_2026',
    19: 'valor_ago_2026',
    20: 'valor_set_2026',
    21: 'valor_out_2026',
    22: 'valor_nov_2026',
    23: 'valor_dez_2026',
}


def safe_float(v):
    try:
        if v is None or str(v).strip() == '':
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def safe_date(v):
    if isinstance(v, datetime):
        return v.date()
    return None


def run_seed():
    with app.app_context():
        print("[DB] Criando tabelas no banco de dados...")
        db.create_all()

        # 1. Criar usuário admin se não existir
        admin = User.query.filter_by(username=ADMIN_USER).first()
        if not admin:
            print("[DB] Criando usuario padrao 'admin'...")
            admin = User(username='admin', role='editor')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print(f"[OK] Usuario '{ADMIN_USER}' criado com senha '{ADMIN_PASS}'")
        else:
            print(f"[INFO] Usuario '{ADMIN_USER}' ja existe, pulando.")

        # 2. Importar PJs do Excel
        print(f"\n[EXCEL] Lendo arquivo: {EXCEL_PATH}")

        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['PJS ATIVOS']

        imported = 0
        skipped = 0

        for row_idx in range(4, ws.max_row + 1):  # Linha 4 em diante (3 = cabeçalho)
            nome = ws.cell(row_idx, 1).value
            cnpj = ws.cell(row_idx, 2).value

            # Pular linhas sem nome ou CNPJ
            if not nome or not cnpj:
                continue

            nome = str(nome).strip()
            cnpj = str(cnpj).strip()

            # Pular CNPJs inválidos ou com texto de fórmula
            if len(cnpj) < 10 or 'SUM' in cnpj.upper():
                continue

            # Verificar se já existe
            if PjContrato.query.filter_by(cnpj=cnpj).first():
                skipped += 1
                continue

            # Determinar se está ativo (tem algum valor em 2026)
            has_value = any(
                safe_float(ws.cell(row_idx, col).value) > 0
                for col in MONTH_COL_MAP.keys()
            )

            pj = PjContrato(
                nome=nome,
                cnpj=cnpj,
                razao_social=str(ws.cell(row_idx, 3).value or '').strip(),
                data_inicio=safe_date(ws.cell(row_idx, 4).value),
                data_encerramento=safe_date(ws.cell(row_idx, 5).value),
                cargo=str(ws.cell(row_idx, 6).value or '').strip(),
                centro_custo=str(ws.cell(row_idx, 8).value or '').strip(),  # col H = CENTRO DE CUSTO (nome)
                observacoes=str(ws.cell(row_idx, 30).value or '').strip(),
                ativo=has_value,
                valor_2025=safe_float(ws.cell(row_idx, 10).value),
            )

            # Valores mensais 2026
            for col, field in MONTH_COL_MAP.items():
                setattr(pj, field, safe_float(ws.cell(row_idx, col).value))

            # Tratar vlookup/erro no centro de custo
            if 'VERIFICAR' in (pj.centro_custo or '').upper() or pj.centro_custo.startswith('='):
                # Tenta usar coluna G (código CC) como fallback
                pj.centro_custo = str(ws.cell(row_idx, 7).value or '').strip()

            db.session.add(pj)
            imported += 1

        db.session.commit()
        print(f"[OK] {imported} PJs importados, {skipped} ja existiam no banco.")

        # 3. Importar CLTs do Excel (Base.xlsx)
        print(f"\n[EXCEL] Lendo arquivo CLT: {CLT_EXCEL_PATH}")
        if os.path.exists(CLT_EXCEL_PATH):
            print("[INFO] Processando base CLT...")
            try:
                df_clt = pd.read_excel(CLT_EXCEL_PATH, sheet_name='base', header=None, skiprows=1)
                
                df_clt_clean = pd.DataFrame({
                    'centro_custo': df_clt[3].astype(str).str.strip(),
                    'matricula': df_clt[4].astype(str).str.strip(),
                    'nome': df_clt[5].astype(str).str.strip(),
                    'situacao': pd.to_numeric(df_clt[34], errors='coerce').fillna(0),
                    'custo_total': pd.to_numeric(df_clt[35], errors='coerce').fillna(0),
                    'total_rem': pd.to_numeric(df_clt[13], errors='coerce').fillna(0),
                    'mes': df_clt[36],
                    'tipo_contrato': 'CLT'
                })
                
                df_clt_clean['mes'] = pd.to_datetime(df_clt_clean['mes'], errors='coerce')
                df_clt_clean.dropna(subset=['matricula', 'mes'], inplace=True)
                df_clt_clean = df_clt_clean[df_clt_clean['centro_custo'] != 'nan']
                
                # Check if there's already data to avoid duplicating
                existing_clt_count = CltRegistro.query.count()
                if existing_clt_count > 0:
                    print(f"[INFO] Ja existem {existing_clt_count} registros CLT no banco. Limpando tabela para reimportar...")
                    db.session.query(CltRegistro).delete()
                    db.session.commit()

                print("[DB] Inserindo registros CLT no banco de dados...")
                clt_records = []
                for _, row in df_clt_clean.iterrows():
                    clt_records.append(CltRegistro(
                        matricula=row['matricula'],
                        nome=row['nome'],
                        centro_custo=row['centro_custo'],
                        mes=row['mes'].date(),
                        situacao=int(row['situacao']),
                        custo_total=float(row['custo_total']),
                        total_rem=float(row['total_rem'])
                    ))
                db.session.bulk_save_objects(clt_records)
                db.session.commit()
                print(f"[OK] {len(clt_records)} registros CLT importados.")
                
            except Exception as e:
                print(f"[ERRO] Falha ao importar Base CLT: {e}")
        else:
            print(f"[AVISO] Arquivo {CLT_EXCEL_PATH} nao encontrado. Pulando importacao CLT.")

        # 4. Importar Agrupadores do depara.xlsx
        print(f"\n[EXCEL] Lendo arquivo de agrupadores: {DEPARA_EXCEL_PATH}")
        if os.path.exists(DEPARA_EXCEL_PATH):
            try:
                df_depara = pd.read_excel(DEPARA_EXCEL_PATH)

                # Only import if table is empty (don't overwrite user-defined rules)
                existing_count = AgrupadorCC.query.count()
                if existing_count > 0:
                    print(f"[INFO] Ja existem {existing_count} regras de agrupador no banco. Pulando importacao do depara.xlsx.")
                    print("[INFO] Para reimportar, apague as entradas manualmente no banco primeiro.")
                else:
                    agrupadores_importados = 0

                    # CLT side: columns 0=CLT, 1=AGRUPADOR, 2=DONOS DO PACOTE
                    clt_col = df_depara.columns[0]   # CLT
                    clt_ag1 = df_depara.columns[1]   # AGRUPADOR
                    clt_ag2 = df_depara.columns[2]   # DONOS DO PACOTE

                    for _, row in df_depara.iterrows():
                        cc = str(row[clt_col]).strip() if pd.notna(row[clt_col]) else ''
                        ag1 = str(row[clt_ag1]).strip() if pd.notna(row[clt_ag1]) else ''
                        ag2 = str(row[clt_ag2]).strip() if pd.notna(row[clt_ag2]) else ''
                        if not cc or cc in ('nan', 'None', ''):
                            continue
                        db.session.add(AgrupadorCC(
                            centro_custo=cc,
                            tipo_contrato='CLT',
                            agrupador1=ag1 or None,
                            agrupador2=ag2 or None,
                            agrupador3=None
                        ))
                        agrupadores_importados += 1

                    # PJ side: columns 4=PJ, 5=AGRUPADOR.1, 6=DONOS DO PACOTE.1
                    pj_col = df_depara.columns[4]    # PJ
                    pj_ag1 = df_depara.columns[5]    # AGRUPADOR.1
                    pj_ag2 = df_depara.columns[6]    # DONOS DO PACOTE.1

                    for _, row in df_depara.iterrows():
                        cc = str(row[pj_col]).strip() if pd.notna(row[pj_col]) else ''
                        ag1 = str(row[pj_ag1]).strip() if pd.notna(row[pj_ag1]) else ''
                        ag2 = str(row[pj_ag2]).strip() if pd.notna(row[pj_ag2]) else ''
                        if not cc or cc in ('nan', 'None', ''):
                            continue
                        db.session.add(AgrupadorCC(
                            centro_custo=cc,
                            tipo_contrato='PJ',
                            agrupador1=ag1 or None,
                            agrupador2=ag2 or None,
                            agrupador3=None
                        ))
                        agrupadores_importados += 1

                    db.session.commit()
                    print(f"[OK] {agrupadores_importados} regras de agrupador importadas do depara.xlsx.")

            except Exception as e:
                print(f"[ERRO] Falha ao importar depara.xlsx: {e}")
        else:
            print(f"[AVISO] Arquivo {DEPARA_EXCEL_PATH} nao encontrado. Pulando importacao de agrupadores.")

        print("\n[OK] Seed concluido! Banco de dados pronto.")
        print(f"\n>> Acesse: http://localhost:5156/login")
        print(f"   Usuario: {ADMIN_USER}")
        print(f"   Senha:   {ADMIN_PASS}")



if __name__ == '__main__':
    run_seed()
